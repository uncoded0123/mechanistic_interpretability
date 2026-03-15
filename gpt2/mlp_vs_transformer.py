import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

# --- Colors ---
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
input_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
output_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
predict_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
bold = Font(bold=True, size=11)
normal = Font(size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_range(ws, row, col, value, fill=None, font=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill: cell.fill = fill
    if font: cell.font = font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')
    return cell

# ============================================================
# TAB 1: MLP
# ============================================================
ws1 = wb.active
ws1.title = "MLP"

# Title
ws1.merge_cells('A1:F1')
style_range(ws1, 1, 1, "MLP: Each row processed INDEPENDENTLY (no row talks to another)", font=Font(bold=True, size=14))

# --- Input Data ---
style_range(ws1, 3, 1, "INPUT DATA (weather spreadsheet)", font=bold)
headers = ["Day", "Temp°F", "Humidity%", "Wind mph", "Pressure hPa", "→ Predict: Rain?"]
for i, h in enumerate(headers):
    fill = header_fill if i < 5 else predict_fill
    style_range(ws1, 4, i+1, h, fill=fill, font=header_font)

data = [
    ["Mon", 72, 85, 10, 1010],
    ["Tue", 68, 90, 15, 1005],
    ["Wed", 75, 60, 5,  1020],
    ["Thu", 65, 95, 20, 1000],
]
rain = ["Yes", "Yes", "No", "?"]

for r, (row_data, pred) in enumerate(zip(data, rain)):
    for c, val in enumerate(row_data):
        style_range(ws1, 5+r, c+1, val, fill=input_fill, font=normal)
    style_range(ws1, 5+r, 6, pred, fill=predict_fill, font=bold)

# Explanation
style_range(ws1, 10, 1, "HOW MLP WORKS:", font=bold)
ws1.merge_cells('A11:F11')
style_range(ws1, 11, 1, "Each row goes through MLP alone. Mon's features → predict Mon. Tue's features → predict Tue.", font=normal)
ws1.merge_cells('A12:F12')
style_range(ws1, 12, 1, "Row for Thu: MLP sees [65, 95, 20, 1000] → outputs 'Yes' (rain). No info from Mon/Tue/Wed used.", font=normal)
ws1.merge_cells('A13:F13')
style_range(ws1, 13, 1, "LIMITATION: Can't learn 'it rained 3 days in a row so drought unlikely' — no cross-row info.", font=Font(bold=True, size=11, color="FF0000"))

# --- MLP internals ---
style_range(ws1, 15, 1, "MLP INTERNALS (Thu row only):", font=bold)
mlp_headers = ["Step", "Values", "What happens"]
for i, h in enumerate(mlp_headers):
    style_range(ws1, 16, i+1, h, fill=header_fill, font=header_font)

mlp_steps = [
    ["Input", "[65, 95, 20, 1000]", "Raw features for Thu"],
    ["Hidden layer 1", "[0.8, 0.2, 0.9, 0.1]", "Weighted sum + ReLU (learns combos)"],
    ["Hidden layer 2", "[0.7, 0.6]", "Further compression"],
    ["Output", "[0.85]", "0.85 > 0.5 → Yes (rain)"],
]
for r, row in enumerate(mlp_steps):
    for c, val in enumerate(row):
        style_range(ws1, 17+r, c+1, val, fill=input_fill, font=normal)

# Column widths
for col_letter, width in [('A',12),('B',22),('C',22),('D',18),('E',18),('F',20)]:
    ws1.column_dimensions[col_letter].width = width

# ============================================================
# TAB 2: TRANSFORMER
# ============================================================
ws2 = wb.create_sheet("Transformer")

ws2.merge_cells('A1:G1')
style_range(ws2, 1, 1, "TRANSFORMER: Rows TALK to each other via Attention, then MLP processes each row", font=Font(bold=True, size=14))

# --- Input ---
style_range(ws2, 3, 1, "INPUT (same weather data)", font=bold)
t_headers = ["Day", "Temp°F", "Humidity%", "Wind mph", "Pressure hPa"]
for i, h in enumerate(t_headers):
    style_range(ws2, 4, i+1, h, fill=header_fill, font=header_font)

for r, row_data in enumerate(data):
    for c, val in enumerate(row_data):
        style_range(ws2, 5+r, c+1, val, fill=input_fill, font=normal)

# --- STEP 1: Attention ---
attn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
style_range(ws2, 10, 1, "STEP 1: ATTENTION (rows share info)", font=bold)

ws2.merge_cells('A11:G11')
style_range(ws2, 11, 1, "Thu asks: 'Which past days are similar to me?' → Finds Tue (also high humidity, low pressure)", font=normal)

attn_headers = ["Day", "Attends to", "Weight", "Info pulled"]
for i, h in enumerate(attn_headers):
    style_range(ws2, 12, i+1, h, fill=header_fill, font=header_font)

attn_data = [
    ["Thu", "Mon", "0.1", "Low attention (different weather)"],
    ["Thu", "Tue", "0.7", "HIGH attention (similar: humid, low pressure, rained)"],
    ["Thu", "Wed", "0.2", "Low attention (sunny day, not similar)"],
]
for r, row in enumerate(attn_data):
    for c, val in enumerate(row):
        style_range(ws2, 13+r, c+1, val, fill=attn_fill, font=normal)

ws2.merge_cells('A16:G16')
style_range(ws2, 16, 1, "Result: Thu's row is now ENRICHED with pattern info from similar past days", font=Font(bold=True, size=11, color="006600"))

# --- STEP 1 result: Updated row ---
style_range(ws2, 18, 1, "Thu row AFTER attention:", font=bold)
updated_headers = ["Temp°F", "Humidity%", "Wind mph", "Pressure hPa", "+ Past rain pattern", "+ Trend info"]
for i, h in enumerate(updated_headers):
    style_range(ws2, 19, i+1, h, fill=header_fill, font=header_font)

updated_vals = [65, 95, 20, 1000, "Tue rained (similar)", "Pressure dropping"]
for i, val in enumerate(updated_vals):
    style_range(ws2, 20, i+1, val, fill=output_fill, font=normal)

# --- STEP 2: MLP ---
style_range(ws2, 22, 1, "STEP 2: MLP (same as MLP tab, but now has MORE info)", font=bold)
ws2.merge_cells('A23:G23')
style_range(ws2, 23, 1, "MLP input: [65, 95, 20, 1000] + attention enrichment → 'Yes' rain (more confident than MLP alone)", font=normal)

# --- Final prediction ---
style_range(ws2, 25, 1, "FINAL OUTPUT:", font=bold)
final_headers = ["Day", "MLP-only prediction", "Transformer prediction", "Why transformer is better"]
for i, h in enumerate(final_headers):
    style_range(ws2, 26, i+1, h, fill=header_fill, font=header_font)

final_data = [
    ["Thu", "Rain 85%", "Rain 95%", "Saw Tue (similar day) also rained"],
]
for r, row in enumerate(final_data):
    for c, val in enumerate(row):
        style_range(ws2, 27+r, c+1, val, fill=predict_fill, font=bold)

ws2.merge_cells('A29:G29')
style_range(ws2, 29, 1, "KEY: Attention lets rows share info → MLP makes better predictions with richer input", font=Font(bold=True, size=13, color="4472C4"))

for col_letter, width in [('A',14),('B',24),('C',22),('D',24),('E',22),('F',22),('G',22)]:
    ws2.column_dimensions[col_letter].width = width

# Save
path = "/home/cody/Documents/Anthropic/mechanistic_interpretability/gpt2/mlp_vs_transformer.xlsx"
wb.save(path)
print(f"Saved to {path}")
